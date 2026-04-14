"""
iris/local_rag.py

Fast local RAG — reads documents from the knowledge folder directly.
No Supabase. No embeddings. No network calls. Pure offline.

Strategy: keyword-overlap scoring (BM25-lite) for near-instant retrieval.
  - Chunks are built once and cached in memory (re-read only when files change).
  - Scoring is 100% in-process Python — ~1–5 ms for 200 chunks.
  - Works without nomic-embed-text installed.

Supported file types: .txt  .md  .pdf (text-layer)  .docx  .xlsx

Usage:
    from .local_rag import search_local_kb, local_kb_status

    result = search_local_kb("what is the TDI byte value for 8-stage?", top_k=3)
    print(result["context_text"])

Environment / config:
    IRIS_KNOWLEDGE_FOLDER  — overrides the default  iris/knowledge/  path
"""

from __future__ import annotations

import os
import re
import math
import time
import hashlib
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ── Knowledge folder path ─────────────────────────────────────────────────────

def _knowledge_dir() -> Path:
    custom = os.environ.get("IRIS_KNOWLEDGE_FOLDER", "")
    if custom and Path(custom).is_dir():
        return Path(custom)
    return Path(__file__).parent / "knowledge"


# ── Chunk cache ───────────────────────────────────────────────────────────────

_CHUNK_CACHE: List[Dict] = []          # [{text, source, heading, file_hash}, ...]
_CACHE_META:  Dict[str, str] = {}      # filename → md5  (change detection)
_LAST_BUILD:  float = 0.0
_CACHE_TTL:   float = 60.0            # seconds before we re-check for new files

CHUNK_SIZE    = 500   # characters
CHUNK_OVERLAP = 80


# ── Text extraction helpers ───────────────────────────────────────────────────

def _extract_txt(path: Path) -> str:
    try:
        return path.read_text(errors="ignore")
    except Exception:
        return ""


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
        r = PdfReader(str(path))
        parts = []
        for page in r.pages:
            t = page.extract_text() or ""
            if t.strip():
                parts.append(t)
        return "\n".join(parts)
    except ImportError:
        pass
    try:
        import subprocess
        result = subprocess.run(
            ["pdftotext", str(path), "-"],
            capture_output=True, text=True, timeout=30
        )
        return result.stdout
    except Exception:
        return ""


def _extract_docx(path: Path) -> str:
    try:
        import docx
        doc = docx.Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception:
        return ""


def _extract_xlsx(path: Path) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            parts.append(f"[Sheet: {sheet}]")
            for row in ws.iter_rows(values_only=True):
                row_text = "  ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    parts.append(row_text)
        return "\n".join(parts)
    except Exception:
        return ""


def _extract(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in (".txt", ".md"):
        return _extract_txt(path)
    if ext == ".pdf":
        return _extract_pdf(path)
    if ext == ".docx":
        return _extract_docx(path)
    if ext in (".xlsx", ".xlsm"):
        return _extract_xlsx(path)
    return ""


# ── Chunker ───────────────────────────────────────────────────────────────────

_RE_HEADING = re.compile(r"^(#{1,4} .+|[A-Z][^\n]{0,60})$", re.M)


def _chunk_text(text: str, source: str) -> List[Dict]:
    """Split text into overlapping chunks, preserving headings as metadata."""
    chunks  = []
    current_heading = ""
    pos = 0
    n   = len(text)

    while pos < n:
        end   = min(pos + CHUNK_SIZE, n)
        chunk = text[pos:end]

        # Detect heading in this chunk
        for m in _RE_HEADING.finditer(chunk):
            current_heading = m.group(0).strip("#").strip()

        chunk_text = chunk.strip()
        if len(chunk_text) > 30:
            chunks.append({
                "text":    chunk_text,
                "source":  source,
                "heading": current_heading,
            })

        step = CHUNK_SIZE - CHUNK_OVERLAP
        pos += step

    return chunks


# ── Cache builder ─────────────────────────────────────────────────────────────

def _md5(path: Path) -> str:
    try:
        h = hashlib.md5()
        with open(path, "rb") as f:
            for block in iter(lambda: f.read(65536), b""):
                h.update(block)
        return h.hexdigest()
    except Exception:
        return ""


def _build_cache(force: bool = False) -> List[Dict]:
    global _CHUNK_CACHE, _CACHE_META, _LAST_BUILD

    now = time.time()
    if not force and _CHUNK_CACHE and (now - _LAST_BUILD) < _CACHE_TTL:
        return _CHUNK_CACHE

    kdir = _knowledge_dir()
    if not kdir.is_dir():
        _CHUNK_CACHE = []
        _LAST_BUILD  = now
        return _CHUNK_CACHE

    SUPPORTED = {".txt", ".md", ".pdf", ".docx", ".xlsx", ".xlsm"}
    all_chunks: List[Dict] = []
    new_meta: Dict[str, str] = {}

    for path in sorted(kdir.rglob("*")):
        if path.suffix.lower() not in SUPPORTED:
            continue
        if path.name.startswith(".") or path.name.startswith("_"):
            continue

        md5 = _md5(path)
        new_meta[path.name] = md5

        # Re-use existing chunks for unchanged files
        if not force and md5 == _CACHE_META.get(path.name):
            existing = [c for c in _CHUNK_CACHE if c["source"] == path.name]
            if existing:
                all_chunks.extend(existing)
                continue

        text = _extract(path)
        if text.strip():
            all_chunks.extend(_chunk_text(text, path.name))

    _CHUNK_CACHE = all_chunks
    _CACHE_META  = new_meta
    _LAST_BUILD  = now
    return _CHUNK_CACHE


# ── BM25-lite scorer ──────────────────────────────────────────────────────────

_RE_WORD = re.compile(r"\b\w+\b")

# Common words to ignore
_STOPWORDS = {
    "the", "a", "an", "is", "it", "in", "on", "of", "to", "for",
    "and", "or", "but", "with", "this", "that", "be", "are", "was",
    "as", "at", "by", "from", "have", "has", "had", "not", "do",
    "does", "did", "will", "would", "could", "should", "may", "might",
    "its", "if", "so", "no", "can", "all", "more", "what", "which",
    "how", "when", "where", "there", "their", "they", "we", "you",
    "i", "my", "your", "our", "any", "each", "into", "about",
}


def _tokenise(text: str) -> List[str]:
    return [
        w.lower() for w in _RE_WORD.findall(text)
        if w.lower() not in _STOPWORDS and len(w) > 1
    ]


def _score_chunk(chunk_tokens: List[str], query_tokens: List[str],
                 idf: Dict[str, float], avgdl: float, k1: float = 1.5, b: float = 0.75) -> float:
    """BM25 score — fast, in-process."""
    dl = len(chunk_tokens)
    freq: Dict[str, int] = {}
    for t in chunk_tokens:
        freq[t] = freq.get(t, 0) + 1

    score = 0.0
    for qt in set(query_tokens):
        if qt not in idf:
            continue
        f  = freq.get(qt, 0)
        tf = (f * (k1 + 1)) / (f + k1 * (1 - b + b * dl / max(avgdl, 1)))
        score += idf[qt] * tf

    return score


def _compute_idf(chunks: List[List[str]]) -> Dict[str, float]:
    N  = len(chunks)
    df: Dict[str, int] = {}
    for toks in chunks:
        for t in set(toks):
            df[t] = df.get(t, 0) + 1
    return {t: math.log((N - n + 0.5) / (n + 0.5) + 1) for t, n in df.items()}


# ── Public API ────────────────────────────────────────────────────────────────

def search_local_kb(
    query:       str,
    top_k:       int   = 4,
    min_score:   float = 0.5,
    file_filter: Optional[str] = None,
) -> Dict:
    """
    Search the local knowledge folder using BM25.

    Returns dict with keys: results, total_found, context_text.
    Never raises — returns empty result on any error.
    """
    if not query or not query.strip():
        return {"query": query, "total_found": 0, "results": [], "context_text": ""}

    chunks = _build_cache()

    if file_filter:
        chunks = [c for c in chunks if file_filter.lower() in c["source"].lower()]

    if not chunks:
        kdir = _knowledge_dir()
        return {
            "query":        query,
            "total_found":  0,
            "results":      [],
            "context_text": (
                f"Knowledge folder is empty or not found at: {kdir}\n"
                "Add .txt, .md, .pdf, .docx, or .xlsx files there."
            ),
        }

    # Tokenise everything
    query_tokens  = _tokenise(query)
    chunk_tokens  = [_tokenise(c["text"]) for c in chunks]
    idf           = _compute_idf(chunk_tokens)
    avgdl         = sum(len(t) for t in chunk_tokens) / max(len(chunk_tokens), 1)

    # Score
    scored: List[Tuple[float, Dict]] = []
    for i, chunk in enumerate(chunks):
        s = _score_chunk(chunk_tokens[i], query_tokens, idf, avgdl)
        if s >= min_score:
            scored.append((s, chunk))

    scored.sort(key=lambda x: x[0], reverse=True)
    top = scored[:top_k]

    if not top:
        return {
            "query":        query,
            "total_found":  0,
            "results":      [],
            "context_text": f"No relevant content found in knowledge folder for: '{query}'",
        }

    results       = []
    context_parts = []
    for score, chunk in top:
        source  = chunk["source"]
        heading = chunk.get("heading", "")
        text    = chunk["text"]

        label = source
        if heading:
            label += f" § {heading}"

        results.append({
            "source":  label,
            "file":    source,
            "heading": heading,
            "text":    text,
            "score":   round(score, 2),
        })
        context_parts.append(f"[{label}]\n{text}")

    return {
        "query":        query,
        "total_found":  len(results),
        "results":      results,
        "context_text": (
            f"Knowledge base — local docs for '{query}':\n\n"
            + "\n\n---\n\n".join(context_parts)
        ),
    }


def local_kb_status() -> Dict:
    """Return a summary of what's in the local knowledge folder."""
    kdir   = _knowledge_dir()
    chunks = _build_cache()

    if not kdir.is_dir():
        return {
            "available": False,
            "folder":    str(kdir),
            "message":   f"Knowledge folder not found: {kdir}",
        }

    # Count files and chunks per source
    files_seen: Dict[str, int] = {}
    for c in chunks:
        files_seen[c["source"]] = files_seen.get(c["source"], 0) + 1

    return {
        "available":    bool(chunks),
        "folder":       str(kdir),
        "file_count":   len(files_seen),
        "chunk_count":  len(chunks),
        "files": [
            {"name": name, "chunks": n}
            for name, n in sorted(files_seen.items())
        ],
    }


def rebuild_cache() -> Dict:
    """Force a full cache rebuild — call after adding new files."""
    chunks = _build_cache(force=True)
    return {"rebuilt": True, "chunk_count": len(chunks)}
