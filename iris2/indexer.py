
from __future__ import annotations

import os
import re
import time
import hashlib
import threading
from pathlib import Path
from typing import Iterator, List, Dict, Set, Tuple, Optional, Callable


# -- Config helpers -----------------------------------------------------------

def _load_config(key: str) -> str:
    cfg_path = Path(__file__).parent / "iris_knowledge.cfg"
    if not cfg_path.exists():
        return ""
    for line in cfg_path.read_text().splitlines():
        line = line.strip()
        if line.startswith(key + "="):
            return line.split("=", 1)[1].strip()
    return ""


def _supabase_client():
    from supabase import create_client
    url = os.environ.get("SUPABASE_URL") or _load_config("supabase_url")
    key = os.environ.get("SUPABASE_SERVICE_KEY") or _load_config("supabase_service_key")
    if not url or not key:
        raise RuntimeError(
            "Supabase credentials not set. "
            "Set SUPABASE_URL and SUPABASE_SERVICE_KEY, or add to iris_knowledge.cfg."
        )
    return create_client(url, key)



from .retrieval import embed_texts as _embed_texts, EMBEDDING_DIM

# -- Chunking constants -------------------------------------------------------

CHUNK_SIZE    = 600
CHUNK_OVERLAP = 80
MIN_CHUNK     = 60
EMBED_BATCH   = 32   # texts per Ollama embedding call

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xls", ".md", ".txt", ".csv"}


# -- File text extractors -----------------------------------------------------

def extract_pdf(path: Path) -> Iterator[Tuple[str, str]]:
    try:
        import pypdf
        reader = pypdf.PdfReader(str(path))
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            if text.strip():
                yield (f"page {i+1}", text)
    except ImportError:
        import PyPDF2
        with open(path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for i, page in enumerate(reader.pages):
                text = page.extract_text() or ""
                if text.strip():
                    yield (f"page {i+1}", text)


def extract_docx(path: Path) -> Iterator[Tuple[str, str]]:
    from docx import Document
    doc = Document(str(path))
    current_heading = ""
    buffer = []
    for para in doc.paragraphs:
        style = para.style.name.lower() if para.style else ""
        text  = para.text.strip()
        if not text:
            continue
        if "heading" in style:
            if buffer:
                yield (current_heading, "\n".join(buffer))
                buffer = []
            current_heading = text
        else:
            buffer.append(text)
    if buffer:
        yield (current_heading, "\n".join(buffer))


def extract_xlsx(path: Path) -> Iterator[Tuple[str, str]]:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            row_text = " | ".join(cells).strip(" |")
            if row_text.replace("|", "").strip():
                rows.append(row_text)
        if rows:
            yield (sheet_name, "\n".join(rows))


def extract_csv(path: Path) -> Iterator[Tuple[str, str]]:
    text = path.read_text(errors="ignore")
    if text.strip():
        yield ("csv", text)


def extract_text(path: Path) -> Iterator[Tuple[str, str]]:
    text = path.read_text(errors="ignore")
    if text.strip():
        yield ("", text)


EXTRACTORS = {
    ".pdf":  extract_pdf,
    ".docx": extract_docx,
    ".xlsx": extract_xlsx,
    ".xls":  extract_xlsx,
    ".csv":  extract_csv,
    ".md":   extract_text,
    ".txt":  extract_text,
}


# -- Chunker ------------------------------------------------------------------

def chunk_text(text: str, size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> List[str]:
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    if len(text) <= size:
        return [text] if len(text) >= MIN_CHUNK else []
    chunks = []
    start  = 0
    while start < len(text):
        end = start + size
        if end >= len(text):
            chunk = text[start:]
        else:
            for pattern in ["\n\n", ". ", "? ", "! ", " "]:
                idx = text.rfind(pattern, start, end)
                if idx > start + size // 2:
                    end = idx + len(pattern)
                    break
            chunk = text[start:end]
        chunk = chunk.strip()
        if len(chunk) >= MIN_CHUNK:
            chunks.append(chunk)
        start = end - overlap
        if start >= len(text):
            break
    return chunks


# -- Embedding ----------------------------------------------------------------

def embed_batch(texts: List[str]) -> List[List[float]]:
    """Embed using local sentence-transformers model. Free, no API key."""
    return _embed_texts(texts)


# -- Main indexer class -------------------------------------------------------

class KnowledgeIndexer:

    def __init__(self, knowledge_folder: str, progress_cb: Optional[Callable] = None):
        self.folder      = Path(knowledge_folder)
        self.progress_cb = progress_cb or print
        self._sb         = None

    def _log(self, msg: str):
        self.progress_cb(msg)

    def _sb_client(self):
        if self._sb is None:
            self._sb = _supabase_client()
        return self._sb

    def _list_supported_files(self) -> List[Path]:
        if not self.folder.exists():
            return []
        return [
            p for p in sorted(self.folder.rglob("*"))
            if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS
        ]

    def _get_indexed_filenames(self) -> Set[str]:
        """Return set of all file_name values currently in Supabase."""
        try:
            result = (
                self._sb_client()
                .table("iris_knowledge")
                .select("file_name")
                .execute()
            )
            return {row["file_name"] for row in (result.data or [])}
        except Exception as e:
            self._log(f"  [warn] Could not query indexed files: {e}")
            return set()

    def _delete_stale_files(self, current_filenames: Set[str]) -> List[str]:
        """
        Auto-delete: remove Supabase chunks for files no longer in the folder.
        Called automatically at the start of index_all().
        """
        indexed = self._get_indexed_filenames()
        stale   = indexed - current_filenames
        if not stale:
            return []
        deleted = []
        for file_name in sorted(stale):
            try:
                self._sb_client() \
                    .table("iris_knowledge") \
                    .delete() \
                    .eq("file_name", file_name) \
                    .execute()
                deleted.append(file_name)
                self._log(f"  Deleted stale: {file_name}")
            except Exception as e:
                self._log(f"  [warn] Could not delete {file_name}: {e}")
        return deleted

    def delete_file_from_index(self, file_name: str) -> Dict:
        """Explicitly delete all chunks for a file. Used by ?forget command."""
        try:
            self._sb_client() \
                .table("iris_knowledge") \
                .delete() \
                .eq("file_name", file_name) \
                .execute()
            return {"deleted": True, "file_name": file_name}
        except Exception as e:
            return {"deleted": False, "error": str(e)}

    def index_all(self, force: bool = False) -> Dict:
        """
        Index all files in the knowledge folder.
        AUTO-DELETE: files removed from folder are deleted from Supabase first.
        Incremental: unchanged files skipped unless force=True.
        """
        if not self.folder.exists():
            return {"error": f"Knowledge folder not found: {self.folder}"}

        files = self._list_supported_files()
        if not files:
            return {"error": f"No supported files found in {self.folder}"}

        # Step 1: auto-delete stale
        current_names = {p.name for p in files}
        deleted = self._delete_stale_files(current_names)
        if deleted:
            self._log(f"Removed {len(deleted)} stale file(s): {', '.join(deleted)}")
        else:
            self._log("No stale files to remove.")

        self._log(f"Indexing {len(files)} file(s) in {self.folder}")

        # Step 2: index each file
        total_chunks = 0
        results      = []
        for file_path in files:
            result = self._index_file(file_path, force=force)
            results.append(result)
            if result.get("chunks_indexed", 0) > 0:
                total_chunks += result["chunks_indexed"]
                self._log(f"  OK {file_path.name}: {result['chunks_indexed']} chunks")
            elif result.get("skipped"):
                self._log(f"  -- {file_path.name}: unchanged (skipped)")
            else:
                self._log(f"  ERR {file_path.name}: {result.get('error', 'unknown')}")

        self._log(
            f"Done. {total_chunks} chunks indexed, {len(deleted)} stale removed."
        )
        return {
            "files_found":   len(files),
            "total_chunks":  total_chunks,
            "stale_deleted": deleted,
            "file_results":  results,
        }

    def index_file(self, file_path: str, force: bool = True) -> Dict:
        return self._index_file(Path(file_path), force=force)

    def _index_file(self, path: Path, force: bool = False) -> Dict:
        ext = path.suffix.lower()
        if ext not in EXTRACTORS:
            return {"file": path.name, "error": f"Unsupported: {ext}"}

        file_hash = self._file_hash(path)

        if not force:
            try:
                existing = (
                    self._sb_client()
                    .table("iris_knowledge")
                    .select("id, metadata")
                    .eq("file_name", path.name)
                    .limit(1)
                    .execute()
                )
                if existing.data:
                    stored_hash = existing.data[0].get("metadata", {}).get("file_hash", "")
                    if stored_hash == file_hash:
                        return {"file": path.name, "skipped": True}
            except Exception:
                pass

        try:
            self._sb_client() \
                .table("iris_knowledge") \
                .delete() \
                .eq("file_name", path.name) \
                .execute()
        except Exception:
            pass

        extractor = EXTRACTORS[ext]
        try:
            sections = list(extractor(path))
        except Exception as e:
            return {"file": path.name, "error": f"Extraction failed: {e}"}

        if not sections:
            return {"file": path.name, "error": "No text extracted"}

        records   = []
        chunk_idx = 0
        for heading_or_page, text in sections:
            for chunk in chunk_text(text):
                records.append({
                    "file_name":     path.name,
                    "file_path":     str(path),
                    "file_type":     ext.lstrip("."),
                    "chunk_index":   chunk_idx,
                    "chunk_text":    chunk,
                    "heading":       heading_or_page if ext in {".docx", ".md", ".txt"} else "",
                    "page_or_sheet": heading_or_page if ext in {".pdf", ".xlsx", ".xls", ".csv"} else "",
                    "char_count":    len(chunk),
                    "metadata":      {"file_hash": file_hash},
                })
                chunk_idx += 1

        if not records:
            return {"file": path.name, "error": "No chunks produced"}

        all_texts      = [r["chunk_text"] for r in records]
        all_embeddings: List[List[float]] = []
        for i in range(0, len(all_texts), EMBED_BATCH):
            batch = all_texts[i: i + EMBED_BATCH]
            try:
                all_embeddings.extend(embed_batch(batch))
            except Exception as e:
                return {"file": path.name, "error": f"Embedding failed at batch {i}: {e}"}

        for record, vec in zip(records, all_embeddings):
            record["embedding"] = vec

        UPSERT_BATCH = 50
        for i in range(0, len(records), UPSERT_BATCH):
            try:
                self._sb_client() \
                    .table("iris_knowledge") \
                    .insert(records[i: i + UPSERT_BATCH]) \
                    .execute()
            except Exception as e:
                return {"file": path.name, "error": f"Supabase insert failed: {e}"}

        return {
            "file":           path.name,
            "file_type":      ext.lstrip("."),
            "chunks_indexed": len(records),
        }

    def index_status(self) -> List[Dict]:
        try:
            result = (
                self._sb_client()
                .table("iris_knowledge_index_status")
                .select("*")
                .execute()
            )
            return result.data or []
        except Exception as e:
            return [{"error": str(e)}]

    @staticmethod
    def _file_hash(path: Path) -> str:
        h = hashlib.md5()
        with open(path, "rb") as f:
            for block in iter(lambda: f.read(65536), b""):
                h.update(block)
        return h.hexdigest()


# -- File watcher (optional) --------------------------------------------------

class _KnowledgeWatcher:
    """
    Watches the knowledge folder for file changes and auto-re-indexes.
    Uses watchdog if available, otherwise polls every 30s.
    """

    def __init__(self, folder: str, progress_cb: Optional[Callable] = None,
                 debounce_sec: float = 3.0):
        self.folder      = folder
        self.progress_cb = progress_cb or print
        self.debounce    = debounce_sec
        self._stop_event = threading.Event()
        self._thread     = None
        self._pending: Optional[threading.Timer] = None
        self._lock       = threading.Lock()

    def start(self):
        self._thread = threading.Thread(
            target=self._run, daemon=True, name="IrisKnowledgeWatcher")
        self._thread.start()
        return self

    def stop(self):
        self._stop_event.set()
        if self._pending:
            self._pending.cancel()

    def _schedule_reindex(self, changed_file: str = ""):
        with self._lock:
            if self._pending:
                self._pending.cancel()
            self._pending = threading.Timer(
                self.debounce, self._do_reindex, args=(changed_file,))
            self._pending.start()

    def _do_reindex(self, changed_file: str = ""):
        try:
            self.progress_cb(
                f"[Iris Watcher] Change detected"
                + (f": {os.path.basename(changed_file)}" if changed_file else "")
                + " — re-indexing…"
            )
            indexer = KnowledgeIndexer(self.folder, progress_cb=self.progress_cb)
            result  = indexer.index_all(force=False)
            if result.get("error"):
                self.progress_cb(f"[Iris Watcher] Error: {result['error']}")
            else:
                n_new = sum(1 for r in result.get("file_results", [])
                            if r.get("chunks_indexed", 0) > 0)
                n_del = len(result.get("stale_deleted", []))
                self.progress_cb(
                    f"[Iris Watcher] Done — {n_new} file(s) updated, {n_del} stale removed.")
        except Exception as e:
            self.progress_cb(f"[Iris Watcher] Error: {e}")

    def _run(self):
        try:
            self._run_watchdog()
        except ImportError:
            self.progress_cb("[Iris Watcher] watchdog not installed — using 30s polling.")
            self._run_polling()

    def _run_watchdog(self):
        from watchdog.observers import Observer
        from watchdog.events import FileSystemEventHandler

        class _Handler(FileSystemEventHandler):
            def __init__(self_, w):
                self_._w = w
            def _relevant(self_, ev):
                return (not ev.is_directory and
                        Path(ev.src_path).suffix.lower() in SUPPORTED_EXTENSIONS)
            def on_modified(self_, ev):
                if self_._relevant(ev): self_._w._schedule_reindex(ev.src_path)
            def on_created(self_, ev):
                if self_._relevant(ev): self_._w._schedule_reindex(ev.src_path)
            def on_deleted(self_, ev):
                if self_._relevant(ev): self_._w._schedule_reindex(ev.src_path)

        observer = Observer()
        observer.schedule(_Handler(self), self.folder, recursive=True)
        observer.start()
        self.progress_cb(f"[Iris Watcher] Watching {self.folder}…")
        try:
            while not self._stop_event.is_set():
                self._stop_event.wait(timeout=1.0)
        finally:
            observer.stop()
            observer.join()

    def _run_polling(self):
        interval = 30

        def snapshot():
            snap = {}
            try:
                for p in Path(self.folder).rglob("*"):
                    if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS:
                        try:
                            st = p.stat()
                            snap[str(p)] = (st.st_mtime, st.st_size)
                        except OSError:
                            pass
            except Exception:
                pass
            return snap

        last = snapshot()
        while not self._stop_event.is_set():
            self._stop_event.wait(timeout=interval)
            if self._stop_event.is_set():
                break
            current = snapshot()
            if current != last:
                last = current
                self._schedule_reindex()


def watch_knowledge_folder(folder: str,
                            progress_cb: Optional[Callable] = None,
                            debounce_sec: float = 3.0) -> "_KnowledgeWatcher":
    """
    Start watching the knowledge folder and auto-re-index on changes.
    Returns watcher — call watcher.stop() to stop.
    Requires watchdog (falls back to 30s polling if not installed).
    """
    return _KnowledgeWatcher(folder, progress_cb=progress_cb,
                              debounce_sec=debounce_sec).start()


# -- Convenience functions called by tools.py ---------------------------------

def run_indexing(knowledge_folder: str,
                 progress_cb: Optional[Callable] = None,
                 force: bool = False) -> Dict:
    """Entry point for tool_index_knowledge()."""
    indexer = KnowledgeIndexer(knowledge_folder, progress_cb=progress_cb)
    return indexer.index_all(force=force)


def delete_from_index(knowledge_folder: str, file_name: str) -> Dict:
    """Delete a specific file. Used by tool_forget_knowledge() / ?forget command."""
    indexer = KnowledgeIndexer(knowledge_folder)
    return indexer.delete_file_from_index(file_name)


def get_index_status(knowledge_folder: str) -> List[Dict]:
    """Return current index status."""
    indexer = KnowledgeIndexer(knowledge_folder)
    return indexer.index_status()